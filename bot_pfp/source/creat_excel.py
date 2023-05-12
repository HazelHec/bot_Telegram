from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment


def creat_densidad(chatId, dt, y, alpha, ac, at, b, atm, bt, A, B, mc, densidadL=0, densidadG=0, densidad=0, z=0, z1=0, z2=0,):
    # libro
    book = Workbook()
    sheet = book.active
    titulos_tabulares = ["Componente", "Masa molar", "Presión Critica",
                         "Temperatura Critica", "Factor acéntrico", "Fracción molar", "alpha", "ac", "at", "b"]
    letras = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L']
    # NOTE - Cambio de tamaño y de color
    range_cells = 'A1:J1'
    for letter in letras:
        sheet.column_dimensions[letter].auto_size = True
        sheet.column_dimensions[letter].alignment = Alignment(
            horizontal='center')

    fill = PatternFill(start_color="F5D7C4",
                       end_color="F5D7C4", fill_type="solid")
    sheet.append(titulos_tabulares)
    for row in sheet[range_cells]:
        for cell in row:
            cell.fill = fill

    # NOTE - Imprime los datos en el excel
    for i in range(len(dt)):
        sheet.append([dt[i]["Componente"], dt[i]["Masa molar"], dt[i]["Presión Critica"], dt[i]
                     ["Temperatura Critica"]+460, dt[i]["Factor acéntrico"], y[i], alpha[i], ac[i], at[i], b[i]])

    # NOTE - Zona de variables con valor único
    sheet["L1"] = "atm"
    sheet["L1"].fill = fill

    sheet["L2"] = atm
    sheet["L3"] = "bt"
    sheet["L3"].fill = fill

    sheet["L4"] = bt
    sheet["L5"] = "A"
    sheet["L5"].fill = fill

    sheet["L6"] = A
    sheet["L7"] = "B"
    sheet["L7"].fill = fill

    sheet["L8"] = B
    sheet["L9"] = "mc"
    sheet["L9"].fill = fill

    sheet["L10"] = mc
    if densidadL == 0:
        sheet["L11"] = "z"
        sheet["L11"].fill = fill

        sheet["L12"] = z
        sheet["L13"] = "densidad"
        sheet["L13"].fill = fill

        sheet["L14"] = densidad
    else:
        sheet["L11"] = "z"
        sheet["L11"].fill = fill
        sheet["L12"] = z
        
        sheet["L13"] = "z1"
        sheet["L13"].fill = fill
        sheet["L14"] = z1
    
        sheet["L17"] = "densidad del liquido"
        sheet["L17"].fill = fill
        sheet["L18"] = densidadL
        
        sheet["L19"] = "densidad del gas"
        sheet["L19"].fill = fill

        sheet["L20"] = densidadG

    book.save(
        f"./resource/datos{chatId}.xlsx")


def creat_presion(chatId, dt, y, alpha, ac, at, b, atm, bt, A, B, mc, vc, vm, presion):
    # libro
    book = Workbook()
    sheet = book.active
    titulos_tabulares = ["Componente", "Masa molar", "Presión Critica",
                         "Temperatura Critica", "Factor acéntrico", "Fracción molar", "alpha", "ac", "at", "b"]
    letras = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L']
    # NOTE - Cambio de tamaño y de color
    range_cells = 'A1:J1'
    for letter in letras:
        sheet.column_dimensions[letter].auto_size = True
        sheet.column_dimensions[letter].alignment = Alignment(
            horizontal='center')

    fill = PatternFill(start_color="FBFF79",
                       end_color="FBFF79", fill_type="solid")
    sheet.append(titulos_tabulares)
    for row in sheet[range_cells]:
        for cell in row:
            cell.fill = fill

    # NOTE - Imprime los datos en el excel
    for i in range(len(dt)):
        sheet.append([dt[i]["Componente"], dt[i]["Masa molar"], dt[i]["Presión Critica"], dt[i]
                     ["Temperatura Critica"]+460, dt[i]["Factor acéntrico"], y[i], alpha[i], ac[i], at[i], b[i]])

    # NOTE - Zona de variables con valor único
    sheet["L1"] = "atm"
    sheet["L1"].fill = fill

    sheet["L2"] = atm
    sheet["L3"] = "bt"
    sheet["L3"].fill = fill

    sheet["L4"] = bt
    sheet["L5"] = "mc"
    sheet["L5"].fill = fill

    sheet["L6"] = mc
    sheet["L7"] = "volumen de la celda"
    sheet["L7"].fill = fill

    sheet["L8"] = vc
    sheet["L9"] = "vm"
    sheet["L9"].fill = fill

    sheet["L10"] = vm
    sheet["L11"] = "presión"
    sheet["L11"].fill = fill

    sheet["L12"] = presion

    book.save(
        f"./resource/datos{chatId}.xlsx")
